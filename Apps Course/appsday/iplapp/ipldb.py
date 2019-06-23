import django
import click
import openpyxl
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE',"appsday.settings")
django.setup()
from iplapp.models import *

@click.command()
def readFromExcel():
    # sheet = "Worksheet"
    # source = "matches.xlsx"
    # insertion_match(source,sheet)
    sheet = "Worksheet"
    source = "deliveries.xlsx"
    insertion_deliveries(source,sheet)


def insertion_deliveries(source,sheet):
    wb1 = openpyxl.load_workbook(source, sheet)
    sheet1 = wb1.get_sheet_by_name(sheet)
    items = []
    for i in range(1,sheet1.max_row + 1):
        items.append([])
        for j in range(1,sheet1.max_column+1):
            e = sheet1.cell(row=i, column=j)
            items[i - 1].append(e.value)
        print(items[i-1][0])
        if i!=1:
            obj = deliveries(
    mid =match(id=(items[i-1][0]+1664)),
    inning = items[i-1][1],
    batting_team = items[i-1][2],
    bowling_team = items[i-1][3],
    over = items[i-1][4],
    ball =items[i-1][5],
    batsman =items[i-1][6] ,
    non_striker = items[i-1][7],
    bowler = items[i-1][8],
    is_super_over =items[i-1][9],
    wide_runs = items[i-1][10],
    bye_runs =items[i-1][11] ,
    legbye_runs =items[i-1][12],
    noball_runs =items[i-1][13] ,
    penalty_runs =items[i-1][14],
    batsman_runs =items[i-1][15] ,
    extra_runs =items[i-1][16] ,
    total_runs =items[i-1][17] ,
    player_dismissed =items[i-1][18] ,
    dismissal_kind =items[i-1][19] ,
    fielder = items[i-1][20])
            obj.save()


def insertion_match(source,sheet):
    wb1 = openpyxl.load_workbook(source, sheet)
    sheet1 = wb1.get_sheet_by_name(sheet)
    items = []
    for i in range(1,sheet1.max_row + 1):
        items.append([])
        for j in range(1,18):
            print(i,j)
            e = sheet1.cell(row=i, column=j)
            items[i - 1].append(e.value)
        print(items[i-1])
        if i!=1 and items[i-1] and items[i-1][8]!="no result" and items[i-1][10]:
            obj = match(season=items[i-1][1],city=items[i-1][2], team1=items[i-1][4], team2=items[i-1][5],
                    toss_won=items[i-1][6], toss_decision=items[i-1][7], result=items[i-1][8], dl_applied=items[i-1][9],winner =items[i-1][10],
                    win_by_runs=items[i-1][11], win_by_wickets=items[i-1][12], player_of_match=items[i-1][13],
                    venue=items[i-1][14], umpire1=items[i-1][15], umpire2=items[i-1][16])
            obj.save()







readFromExcel()