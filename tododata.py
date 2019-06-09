import django
import click
import openpyxl
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE',"homeproject.settings")
django.setup()
from todoapp.models import *

@click.command()
def readFromExcel():
    # sheet = "Colleges"
    # source = "students.xlsx"
    # citems = insertion(source,sheet)
    # for i in range(1, len(citems)):
    #     obj = college(name=citems[i][0], location=citems[i][2], acronym=citems[i][1], contact=citems[i][3])
    #     obj.save()
    # sheet = "Current"
    # source = "students.xlsx"
    # sitems = insertion(source, sheet)
    # for i in range(1, len(sitems)):
    #     sitems[i][3]=sitems[i][3].lower()
    #     obj = student(name=sitems[i][0], dob='2000-10-10', email=sitems[i][2], college_id = sitems[i][1],db_folder = sitems[i][3])
    #     obj.save()
    # sheet = "Sheet1"
    # source = "outputexcel.xlsx"
    # mitems = insertion(source, sheet)
    # for i in range(len(mitems)):
    #     temp=mitems[i][0]
    #     temp=temp[temp.find("_")+1:]
    #     temp = temp[temp.find("_") + 1:]
    #     temp = temp[:temp.find("_")]
    #     mitems[i][0]=temp.lower()
    #
    # for i in range(1, len(mitems)):
    #     print(mitems[i][0])
    #     obj = MockTest1(problem1=mitems[i][1],problem2=mitems[i][2],problem3=mitems[i][3],problem4=mitems[i][4],total=mitems[i][5], student_id = mitems[i][0])
    #     obj.save()
    #uncomment when required
    source = "students.xlsx"
    sheet = "Deletions"
    sitems = insertion(source, sheet)
    for i in range(1, len(sitems)):
        sitems[i][3]=sitems[i][3].lower()
        obj = student(name=sitems[i][0], email=sitems[i][2], college_id = sitems[i][1],db_folder = sitems[i][3])
        obj.save()
    c = college.objects.all()
    print(c)

def insertion(source,sheet):
    wb1 = openpyxl.load_workbook(source, sheet)
    sheet1 = wb1.get_sheet_by_name(sheet)
    items = []
    for i in range(1, sheet1.max_row + 1):
        items.append([])
    for i in range(1, sheet1.max_row + 1):
        for j in range(1, sheet1.max_column + 1):
            e = sheet1.cell(row=i, column=j)
            items[i - 1].append(e.value)
    return items


readFromExcel()